#!/bin/env python3

import numpy as np
import openpyxl as pyxl
import re
import math
from pathlib import Path

script_dir = Path(__file__).parent.resolve()

optics='16FEB24'
vfile=['FACET2e_value.echo','FACET2p_value.echo']

outdir='oracle_upload'
xfile='FACET2_'+optics+'.xls'

print(' ')
print('   ===============================================')
print('           FACET2 Excel File Generation')
print('   ===============================================')
print(' ')

stepnum=0;

Er=5.10998918e5     # electron rest mass (eV) ... XAL value
clight=2.99792458e8 # speed of light (m/s) ... XAL value
charge=-1           # sign of electron charge ... XAL value
pi=3.141592653
rad2deg=180/pi      # degrees per radian
T2kG=10             # kG per Tesla

# ==============================================================================
# hardwired LCLS2sc MAD/XAL stuff
# ------------------------------------------------------------------------------

# file name roots

froot = [
    'FACET2e',     #  1
    'FACET2s',    #  2
    'FACET2p',  #  3
]

seq = []
seq.append({'froot':1,'name':'CATHODE TO DL10',     'beg':'BEGINJ',     'end':'ENDINJ',     'offset':[0,0],  'prev':0,  'suml':0,  'length':0})
seq.append({'froot':1,'name':'DL10 TO L1F',         'beg':'BEGDL10',    'end':'ENDDL10',    'offset':[0,0],  'prev':1,  'suml':0,  'length':0})
seq.append({'froot':1,'name':'L1F TO BC11 CHICANE', 'beg':'BEGL1F',     'end':'BC11CBEG',   'offset':[0,-1], 'prev':2,  'suml':0,  'length':0})
seq.append({'froot':1,'name':'BC11 CHICANE',        'beg':'BC11CBEG',   'end':'BC11CEND',   'offset':[0,0],  'prev':3,  'suml':0,  'length':0})
seq.append({'froot':1,'name':'BC11 CHICANE TO L2F', 'beg':'BC11CEND',   'end':'ENDBC11_2',  'offset':[+1,0], 'prev':4,  'suml':0,  'length':0})
seq.append({'froot':1,'name':'L2F TO BC14 CHICANE', 'beg':'BEGL2F',     'end':'BEGBC14E',   'offset':[0,-1], 'prev':5,  'suml':0,  'length':0})
seq.append({'froot':1,'name':'BC14E CHICANE',       'beg':'BEGBC14E',   'end':'ENDBC14E',   'offset':[0,0],  'prev':6,  'suml':0,  'length':0})
seq.append({'froot':1,'name':'BC14 CHICANE TO L3F', 'beg':'ENDBC14E',   'end':'ENDBC14_2',  'offset':[+1,0], 'prev':7,  'suml':0,  'length':0})
seq.append({'froot':1,'name':'L3F TO SCAV',         'beg':'BEGL3F_1',   'end':'ENDL3F_1',   'offset':[0,0],  'prev':8,  'suml':0,  'length':0})
seq.append({'froot':1,'name':'SCAV TO BC20E',       'beg':'BEGL3F_2',   'end':'ENDL3F_2',   'offset':[0,0],  'prev':9,  'suml':0,  'length':0})
seq.append({'froot':1,'name':'BC20E TO FF20',       'beg':'BEGBC20',    'end':'ENDBC20',    'offset':[0,0],  'prev':10, 'suml':0,  'length':0})
seq.append({'froot':1,'name':'FF20 TO EXPT20',      'beg':'BEGFF20',    'end':'ENDFF20',    'offset':[0,0],  'prev':11, 'suml':0,  'length':0})
seq.append({'froot':1,'name':'EXPT20 TO SPECT20',   'beg':'BEGEXPT20',  'end':'ENDEXPT20',  'offset':[0,0],  'prev':12, 'suml':0,  'length':0})
seq.append({'froot':1,'name':'SPECT20 TO DUMP',     'beg':'BEGSPECT20', 'end':'ENDSPECT20', 'offset':[0,0],  'prev':13, 'suml':0,  'length':0})
# FACET2 scavenger e-
seq.append({'froot':2,'name':'SCAV TO TARGET',      'beg':'BEGSCAV',    'end':'ENDSCAV',    'offset':[0,0],  'prev':9,  'suml':0,  'length':0})
# FACET2 e+
seq.append({'froot':3,'name':'BC14P CHICANE',       'beg':'BEGBC14P',   'end':'ENDBC14P',   'offset':[0,0],  'prev':6,  'suml':0,  'length':0})

seqexcl = []

# ------------------------------------------------------------------------------
# machine areas
area = []
area.append({'name':'INJ',     'beg':'BEGINJ',     'end':'L0AFBEG',    'offset':[0,-1]})
area.append({'name':'L0F',     'beg':'L0AFBEG',    'end':'ENDINJ',     'offset':[0,0]})
area.append({'name':'DL10',    'beg':'BEGDL10',    'end':'ENDDL10',    'offset':[0,0]})
area.append({'name':'L1F',     'beg':'BEGL1F',     'end':'ENDL1F',     'offset':[0,0]})
area.append({'name':'BC11_1',  'beg':'BEGBC11_1',  'end':'ENDBC11_1',  'offset':[0,0]})
area.append({'name':'BC11_2',  'beg':'BEGBC11_2',  'end':'ENDBC11_2',  'offset':[0,0]})
area.append({'name':'L2F',     'beg':'BEGL2F',     'end':'ENDL2F',     'offset':[0,0]})
area.append({'name':'BC14_1',  'beg':'BEGBC14_1',  'end':'ENDBC14_1',  'offset':[0,0]})
area.append({'name':'BC14E',   'beg':'BEGBC14E',   'end':'ENDBC14E',   'offset':[0,0]})
area.append({'name':'BC14_2',  'beg':'BEGBC14_2',  'end':'ENDBC14_2',  'offset':[0,0]})
area.append({'name':'L3F_1',   'beg':'BEGL3F_1',   'end':'ENDL3F_2',   'offset':[0,0]})
area.append({'name':'L3F_2',   'beg':'BEGL3F_2',   'end':'ENDL3F_2',   'offset':[0,0]})
area.append({'name':'BC20',    'beg':'BEGBC20',    'end':'ENDBC20',    'offset':[0,0]})
area.append({'name':'FF20',    'beg':'BEGFF20',    'end':'ENDFF20',    'offset':[0,0]})
area.append({'name':'EXPT20',  'beg':'BEGEXPT20',  'end':'ENDEXPT20',  'offset':[0,0]})
area.append({'name':'SPECT20', 'beg':'BEGSPECT20', 'end':'ENDSPECT20', 'offset':[0,0]})
area.append({'name':'SCAV',    'beg':'BEGSCAV',    'end':'ENDSCAV',    'offset':[0,0]})
area.append({'name':'BC14P',   'beg':'BEGBC14P',   'end':'ENDBC14P',   'offset':[0,0]})

areaexcl = []

from xtffs2mat import xtffs2mat

# ------------------------------------------------------------------------------
# read the MAD output files
stepnum += 1
print('   {}) Read MAD output files ...'.format(stepnum))

K, N, L, P, A, T, E, FDN, coor, S = [], [], [], [], [], [], [], [], [], []
idf, ids, idd = [], [], []  # idf: which MAD output file an element came from
                            # ids: which XAL sequence an element belongs to
                            # idd: ordinal position in MAD output file
nf = 0

for n in range(len(seq)):
    if n in seqexcl:
      continue
    if seq[n]['froot'] != nf:
        nf = seq[n]['froot']
        fname = '{}_survey.tape'.format(froot[nf-1])
        print('Opening file {}'.format(fname))
        titl, tK, tN, tL, tP, tA, tT, tE, tFDN, tcoor, tS = xtffs2mat(fname)

    id1 = tN.index(seq[n]['beg']) + seq[n]['offset'][0]
    id2 = tN.index(seq[n]['end']) + seq[n]['offset'][1]
    id_range = slice(id1, id2 + 1)
    
    K.extend(tK[id_range])
    len_id = len(tK[id_range])
    N.extend(tN[id_range])
    L.extend(tL[id_range])
    P.extend(tP[id_range])
    A.extend(tA[id_range])
    T.extend(tT[id_range])
    E.extend(tE[id_range])
    FDN.extend(tFDN[id_range])
    coor.extend(tcoor[id_range])
    S.extend(tS[id_range])
    idf.extend([nf] * len_id)
    ids.extend([n] * len_id)
    idd.extend(list(range(id1,id2+1)))

Nelem, Nc = len(N), len(N[0])

# set "display S" to linac Z-coordinate
Sd = [x[2] for x in coor]

def intersection(x,y):
    return [v for v in x if v in y]

def strmatch(n_str,N_lst,exact=False):
    if exact:
        return [ix for ix,n_ in enumerate(N_lst) if n_.strip() == n_str.strip()]
    else:
        return [ix for ix,n_ in enumerate(N_lst) if n_.startswith(n_str)]

P2 = []

# initialize some sequence data

for s in seq:
    id1 = N.index(s['beg']) + s['offset'][0]
    id2 = N.index(s['end']) + s['offset'][1]
    s['suml'] = S[id1]
    s['length']=S[id2]-S[id1]

# assign machine areas

ida=[]
for ix,a in enumerate(area):
    id1 = N.index(a['beg']) + a['offset'][0]
    id2 = N.index(a['end']) + a['offset'][1]
    ida.extend([ix]*(id2-id1+1))

# assign area "parent" names

for a in area:
    if a['name'] in ['BC11_1','BC11_2']:
        a['parent']='BD11'
    elif a['name'] in ['BC14_1','BC14E','BC14P','BC14_2']:
        a['parent']='BC14'
    elif a['name'] in ['L3F_1','L3F_2']:
        a['parent']='L3F'
    else:
        a['parent']=a['name']

# kicker/septum group
KSname= ['BKY170']

# read FINT values for SBENs and undulator parameters from a special echo-file
# generated via MAD VALUE commands

C = []
for n in range(len(vfile)):
    fname = vfile[n]
    with open(fname, 'r') as f:
        C.extend(f.read().split())

P2 = np.zeros((Nelem, 2))

idb = [i for i,x in enumerate(K) if x == 'SBEN']
for m in range(0, len(idb), 2):
    na = idb[m]
    nb = idb[m+1]
    name = N[na].strip()
    name = name.split('.')[0]  # remove decoration, if any
    id_ = strmatch(name,C)
    if not id_:
        raise ValueError(f'No FINT for {name}')
    elif len(id_) > 1:
        print('oops')
    id_ = id_[0]
    fint = float(C[id_+6])
    P2[na][0] = fint
    P2[nb][0] = fint

idm = strmatch('MATR',K)
for m in range(0, len(idm), 2):
    n1 = idm[m]
    n2 = idm[m+1]
    name = N[n1].strip()
    Ktxt = f'"{name}_K"'
    Ltxt = f'"{name}_L"'
    idK = strmatch(Ktxt,C)[0]
    idL = strmatch(Ltxt,C)[0]
    undk = float(C[idK+2])
    undl = float(C[idL+2])
    P2[n1, :] = [undl, undk]
    P2[n2, :] = [undl, undk]

# devices between kicker(s) and septum ... shared
anames= ['SCAV'    ,'SCAV'    ,'SCAV'    ]
names = ['BPM19501','BPM19601','BPM19701']
for aname,name in zip(anames,names):
  for jd in strmatch(name,N,True):
    if aname == area(ida(jd)).name:
      N[jd]=name+'?'

# change keyword for gun solenoid correction quads from MULT to QUAD;
# copy T1 into TILT slot
names= ['CQ10121', 'SQ10122']
for name in names:
  id_=strmatch(name, N, True)
  K[id_]='QUAD'
  P[id_,4]=P[id_,6] # T1 -> TILT


# Assign sector names (use empty FDN and FDN1 arrays)

def notnone(val):
  if val is None:
    return ''
  else:
    return val

def read_sector():
  filename = f'{script_dir}/sectors.xlsx'
  wb= pyxl.load_workbook(filename,data_only=True)

  # read worksheet 1 (scS)
  sheet = wb.worksheets[0]
  data = sheet['A4':'J15']
  sect_F2 = []
  for row in data:
    sect_sc.append({
      'name': row[0].value,
      'froot': [row[1].value],
      'BSY': row[2].value,
      'Zbeg': row[4].value,
      'Zend': row[5].value,
      'Nbeg': notnone(row[8].value),
      'Nend': notnone(row[9].value)
    })

  # nf= 1: FACET2e
  # nf= 2: FACET2s
  # nf= 3: FACET2p

  # assign multiple lines to some sectors
  sect_F2[4]['froot'].extend([3])  # S14

  return sect_F2

def set_sector(N, FDN, coor, idf, nf, sector):
  if nf not in sector['froot']:
    return FDN
  Z = [x[2] for x in coor]

  id_ = [i for i,x in enumerate(idf) if x == nf]

  if sector['Nbeg'] == '':
    jd2 = [i for i,x in enumerate(Z) if x > sector['Zbeg']]
    inter1 = intersection(id_,jd2)
    if inter1 == []:
      return FDN
    else:
      inter1 = inter1[0]
  else:
    jd2 = strmatch(sector['Nbeg'],N,True)
    inter1 = intersection(id_,jd2)
    if inter1 != []:
      inter1 = inter1[0]

  if sector['Nend'] == '':
    jd2 = [i for i,x in enumerate(Z) if x < sector['Zend']]
    inter2 = intersection(id_,jd2)
    if inter2 == []:
      return FDN
    else:
      inter2 = inter2[-1]
  else:
    jd2 = strmatch(sector['Nend'],N,True)
    inter2 = intersection(id_,jd2)
    if inter2 != []:
      inter2 = inter2[0]

  if inter1 != [] and inter2 != []:
    for n in range(inter1, inter2 + 1):
      if FDN[n].strip() == '':
        FDN[n] = sector['name']

  return FDN

def assign_sector(N, FDN, coor, idf):
  # NOTE: coordinates are assumed to be in MAD (not SYMBOLS) order

  sect_F2=read_sector()

  # nf= 1: FACET2e
  # nf= 2: FACET2s
  # nf= 3: FACET2p

  # FACET2 LINEs
  for nf in range(1, 4):  # idf values
    for ns, sector in enumerate(sect_F2, 1):
      if nf==3 and ns==5:
        sector['Nbeg'] = 'BEGBC14P'
        sector['Nend'] = 'ENDBC14P'

  FDN=set_sector(N,FDN,coor,idf,nf,sector);

# assign sector names (use empty FDN array)
FDN = assign_sector(N, FDN, coor, idf)

names = ['YC57145','YC57146','BPM19501?','BPM19601?','BPM19701?']
snames = ['EP01','EP01','S19', 'S19', 'S19']
for name, sname in zip(names,snames):
  id_ = strmatch(name, N, True)
  FDN[id_]=sname

# MAD SURVEY coordinates  [x,y,z,theta,phi   ,psi]
# correspond to SolidEdge [z,x,y,roll ,-pitch,yaw]

ic = [2, 0, 1, 5, 4, 3]
coor = np.array(coor)
coor = coor[:, ic]
coor[:, 4] = -coor[:, 4]

# generate ordered list of MAD keywords

tkeyw = sorted(list(set(K)))

MADK = [
    'LCAV', 'SBEN', 'QUAD', 'SEXT', 'SOLE', 'MATR', 'RCOL', 'ECOL', 'SROT',
    'HKIC', 'VKIC', 'MONI', 'WIRE', 'PROF', 'IMON', 'BLMO', 'INST',
    'MARK'
]
OUTK = [
    'LCAV', 'BEND', 'QUAD', 'SEXT', 'SOLE', 'USEG', 'COLL', 'PC  ', 'MARK',
    'XCOR', 'YCOR', 'BPM ', 'WIRE', 'PROF', 'IMON', 'BLMO', 'INST',
    'MARK'
]
XALK = [
    'BNCH', 'BEND', 'QUAD', 'SEXT', 'SOLE', 'USEG', 'COLL', 'ECOL', 'SROT',
    'XCOR', 'YCOR', 'BPM ', 'WIRE', 'PROF', 'TORO', 'BLMO', 'INST',
    'MARK'
]

idmisc = list(range(9, 16))  # non-magnet elements that might have nonzero lengths
keyw = []
keyo = []
keyx = []
jdmisc = []
for n in range(len(MADK)):
    id_ = strmatch(MADK[n], tkeyw)
    if len(id_) > 0:
        keyw.append(MADK[n])
        keyo.append(OUTK[n])
        keyx.append(XALK[n])
        if n in idmisc:
            jdmisc.append(len(keyw) - 1)
idmisc=jdmisc

# hard-wired list of bends that have energy polynomials in the database
Ebend = []  # ['BRB', 'BXSP', 'BYSP', 'BRSP', 'BYSP', 'BX3', 'BY1', 'BY2', 'BYD']

# Keyword Worksheets
# (NOTE: suml, energy, and coord are quoted at element's center)
# ==============================================================================
# LCAV (unsegmented):
# - idf,id,sequence,area,xkey,prim,name,type,dist(m),energy(GeV)
# - length(m),freq(MHz),ampl(MeV),phase(deg),grad(MeV/m),power(1)
# - sdsp(m),suml(m),coord(m,rad)
# - suml1(m),coord1(m,rad),suml2(m),coord2(m,rad)
# SBEN (unsplit):
# - idf,id,sequence,area,xkey,prim,name,type,dist(m),energy(GeV)
# - zleng(m),leng(m),gap(m),fint(1),tilt(deg),ang(deg),e1(deg),e2(deg),
#   BL(kG-m),B(T),k1(1/m^2),GL(kG),G(T/m),scale(name,value),polarity
# - sdsp(m),suml(m),coord(m,rad),mcoord(m,rad)
# - suml1(m),coord1(m,rad),mcoord1(m,rad),suml2(m),coord2(m,rad),mcoord2(m,rad)
# QUAD (unsplit):
# - idf,id,sequence,area,xkey,prim,name,type,dist(m),energy(GeV)
# - leng(m),bore(m),tilt(deg),k1(1/m^2),GL(kG),G(T/m),scale(name,value),polarity
# - sdsp(m),suml(m),coord(m,rad),mcoord(m,rad)
# - suml1(m),coord1(m,rad),suml2(m),coord2(m,rad)
# SEXT (unsplit):
# - idf,id,sequence,area,xkey,prim,name,type,dist(m),energy(GeV)
# - leng(m),bore(m),tilt(deg),k2(1/m^3),G'L(kG/m),G'(T/m^2),scale(name,value),
#   polarity
# - sdsp(m),suml(m),coord(m,rad)
# - suml1(m),coord1(m,rad),suml2(m),coord2(m,rad)
# SOLE (unsplit):
# - idf,id,sequence,area,xkey,prim,name,type,dist(m),energy(GeV)
# - leng(m),bore(m),ks(1/m),BL(kG-m),B(T),scale(name,value),polarity
# - sdsp(m),suml(m),coord(m,rad)
# - suml1(m),coord1(m,rad),suml2(m),coord2(m,rad)
# MATR (unsplit):
# - idf,id,sequence,area,xkey,prim,name,type,dist(m),energy(GeV)
# - leng(m),lambda(m),k(1)
# - sdsp(m),suml(m),coord(m,rad)
# - suml1(m),coord1(m,rad),suml2(m),coord2(m,rad)
# RCOL,ECOL:
# - idf,id,sequence,area,xkey,prim,name,type,dist(m),energy(GeV)
# - leng(m),xsize(m),ysize(m)
# - sdsp(m),suml(m),coord(m,rad)
# - suml1(m),coord1(m,rad),suml2(m),coord2(m,rad)
# SROT:
# - idf,id,sequence,area,xkey,prim,name,type,dist(m),energy(GeV)
# - leng(m),ang(deg)
# - sdsp(m),suml(m),coord(m,rad)
# - suml1(m),coord1(m,rad),suml2(m),coord2(m,rad)
# HKIC,VKIC,MONI,WIRE,PROF,IMON,BLMO,INST:
# - idf,id,sequence,area,xkey,prim,name,type,dist(m),energy(GeV)
# - leng(m)
# - sdsp(m),suml(m),coord(m,rad)
# - suml1(m),coord1(m,rad),suml2(m),coord2(m,rad)
# MARK:
# - idf,id,sequence,area,xkey,prim,name,type,dist(m),energy(GeV)
# - sdsp(m),suml(m),coord(m,rad)
# - suml1(m),coord1(m,rad),suml2(m),coord2(m,rad)

# Assumptions
# ==============================================================================
# - LCAVs may be segmented
# - the first 6 characters of an LCAV's name associate it with it's parent
# - all SBENs are split into two pieces
# - the last character of an SBEN's name differentiates it's pieces
# - all QUADs, SEXTs, and MATRs are split in half
# - all other keyword types are not (necessarily) split
# - SBENs with abs(ang)<amin will have ang set to zero
# - SBENs or QUADs with abs(k1)<kmin will have k1 set to zero
# - SOLEs with abs(ks)<kmin will have ks set to zero
# - SROTs with abs(ang)<amin will have ang set to zero

amin = 1e-9
kmin = 1e-6

# process by keyword

nHKIC = 0
nVKIC = 0
nMONI = 0
nWIRE = 0
nPROF = 0
nIMON = 0
nBLMO = 0
nINST = 0

def slicer(N,ix):
    return [N[i] for i in ix]

#HEAD
Sd = np.array(Sd)
S = np.array(S)
L = np.array(L)
P = np.array(P)
P2 = np.array(P2)
E = np.array(E)
seq = np.array(seq)
A = np.array(A)

for kwn,kxn,kon in zip(keyw,keyx,keyo):
    id_ = strmatch(kwn,K)
    if kwn == 'LCAV':
        # create list of unique names that will allow unsplitting
        name = slicer(N,id_)
        name = list(dict.fromkeys(name))  # unique stable
        LCAV = []
        nLCAV = 0

        for mname in name:
            if mname == 'K19_5X' or mname == 'K19_6X':
              continue
            else:
              nLCAV += 1

            id_ = strmatch(mname,N)
            id1 = id_[0]  # first segment
            ide = [id1-1, id_[-1]]  # [entrance, exit]
            sdsp = np.mean(Sd[ide])  # m (beam center)
            suml = np.mean(S[ide])  # m (beam center)
            dist = suml - seq[ids[id1]]['suml']  # m (sequence start to beam center)
            energy = np.mean(E[ide])  # GeV (beam center)
            leng = np.sum(L[id_])  # m
            freq = P[id1, 4]  # MHz
            ampl = np.sum(P[id_, 5])  # MeV
            phase = P[id1, 6]  # rad/2pi
            grad = ampl / leng  # MeV/m

            if re.match(r'K\d\d_\d[ABCD]', mname[0:6]):  # i.e. K27_3D
                id_ = strmatch(mname[0:5],N)
                grad0 = np.min(P[id_, 5] / L[id_])
                if grad0 == 0:
                    power = 0.25 # KLYS power fraction (1)
                else:
                    power = 0.25 * round((grad / grad0) ** 2)  # KLYS power fraction (1)
            else:
                power = 1

            eloss = P[id1,8]  # V/C
            coorc = np.mean(coor[ide, :], axis=0)  # m,rad (beam center)
            nLCAV += 1

            if mname[4:6] == '__':
              mname_ = mname[0:4]
            elif mname[0:2] == 'TC':
              mname_ = N[id1].rstrip()
            else:
              mname_ = mname;

            LCAV.append({
                'idf': idf[id1],
                'id': idd[id1],
                'seq': seq[ids[id1]]['name'],
                'area': area[ida[id1]]['name'],
                'parent': area[ida[id1]]['parent'],
                'sector': FDN[id1].strip(),
                'xkey': kxn,
                'prim': kon,
                'name': mname_,
                'type': T[id1].strip(),
                'dist': dist,
                'energy': energy,
                'leng': leng,
                'freq': freq,
                'ampl': ampl,
                'phase': 360 * phase,  # deg
                'grad': grad,
                'power': power,
                'sdsp': sdsp,
                'suml': suml,
                **{f'c{k+1}': coorc[k] for k in range(6)}
            })
    elif kwn == 'SBEN':
        name = [N[i] for i in id_]
        SBEN = []
        Nelm = len(id_) // 2
        for m in range(Nelm):
            mname1 = name[2 * m].strip()
            mname2 = name[2 * m + 1].strip()
            mname = mname1[:-1]  # remove last character from name
            if mname in ['Q19501' ,'Q19601' ,'Q19701' ,'BCX141720' ,'BCX141883']:
              continue
            else:
              nSBEN=nSBEN+1

            id_ = [strmatch(mname1,N,True),strmatch(mname2,N,True)]
            id1 = id_[0][0]  # first piece (beam center)
            idi = id1 - 1  # beam in
            ido = id_[1][-1]  # beam out
            sdsp = Sd[id1]  # m
            suml = S[id1]  # m
            dist = suml - seq[ids[id1]]['suml']  # m
            energy = E[id1]  # GeV
            leng = np.sum(L[id_])  # m
            gap = 2 * A[id1]  # m
            fint = P2[id1, 0]  # m
            tilt = P[id1, 3]  # rad
            ang = np.sum(P[id_, 0])  # rad
            if abs(ang) < amin:
                ang = 0
                e1 = 0
                e2 = 0
            else:
                e1 = P[id1, 4]  # rad
                e2 = P[ido, 5]  # rad
            EeV = 1e9 * energy  # eV
            brho = np.sqrt(EeV ** 2 - Er ** 2) / clight  # T-m
            BL = brho * ang  # T-m
            B = BL / leng  # T
            k1 = P[id1, 1]  # 1/m^2
            if abs(k1) < kmin:
                k1 = 0
            G = brho * k1  # T/m
            GL = G * leng  # T
            if mname[:3] in Ebend:
                sname = 'GeV2T'
                sval = brho * abs(ang) / (leng * energy)
            else:
                sname = 'kG2T_Bdl2B'
                sval = 1 / (leng * T2kG)
            polarity = -np.sign(ang + np.finfo(float).eps)  # add eps so that sign=1 when ang=0
            coori = np.copy(coor[idi, :])  # m,rad
            coorc = np.copy(coor[id1, :])  # m,rad
            cooro = np.copy(coor[ido, :])  # m,rad
            coorm = np.zeros(coorc.shape)  # m,rad (magnet steel center)
            if mname in KSname:
                jd = strmatch(f'D{mname}',N)
                if len(jd) != 2:
                    raise ValueError(f'{mname} not split?')
                zleng = np.sum(L[jd])  # m
                coorm = np.copy(coor[jd[0], :])  
                coorm[3] = 0
            else:
                chicane1 = (e1 == 0) & (e2 != 0)
                chicane2 = (e1 != 0) & (e2 == 0)
                if chicane1 | chicane2:
                    zleng = leng * np.sinc(ang/np.pi)  # m
                    coorm[:3] = np.mean([coori[:3], cooro[:3]], axis=0)
                    if chicane1:
                        coorm[3:6] = np.copy(coori[3:6])
                    else:
                        coorm[3:6] = np.copy(cooro[3:6])
                else:
                    zleng = leng * np.sinc(ang / 2 / np.pi)  # m
                    coorm[:3] = (coori[:3] + cooro[:3] + 2 * coorc[:3]) / 4
                    coorm[3:6] = np.copy(coorc[3:6])
            SBEN.append({
                'idf': idf[id1],
                'id': idd[id1],
                'seq': seq[ids[id1]]['name'],
                'area': area[ida[id1]]['name'],
                'parent': pname,
                'sector': FDN[id1].strip(),
                'xkey': f'X{kxn}' if tilt == 0 else f'Y{kxn}' if abs(np.cos(tilt)) < 1e-9 else f'R{kxn}',
                'prim': kon,
                'name': mname,
                'type': T[id1].strip(),
                'dist': dist,
                'energy': energy,
                'zleng': zleng,
                'leng': leng,
                'gap': 2 * A[id1],  # m
                'fint': fint,
                'tilt': np.rad2deg(tilt),  # deg
                'ang': np.rad2deg(ang),  # deg
                'e1': np.rad2deg(e1),  # deg
                'e2': np.rad2deg(e2),  # deg
                'BL': T2kG * BL,  # kG-m
                'B': charge * B,
                'k1': k1,
                'GL': T2kG * GL,  # kG
                'G': charge * G,
                'sname': sname,
                'sval': sval,
                'polarity': polarity,
                'sdsp': sdsp,
                'suml': suml,
                **{f'c{k+1}': coorc[k] for k in range(6)},
                **{f'm{k+1}': coorm[k] for k in range(6)}
            })
    elif kwn == 'QUAD':
        name = list(dict.fromkeys([N[i] for i in id_]))
        QUAD = []
        for mname in name:
            id_ = strmatch(mname,N,True)
            id1 = id_[0]  # first segment (beam center)
            sdsp = Sd[id1]  # m
            suml = S[id1]  # m
            dist = suml - seq[ids[id1]]['suml']  # m
            energy = E[id1]  # GeV
            leng = np.sum(L[id_])  # m
            bore = 2 * A[id1]  # m
            tilt = P[id1, 3]  # rad
            k1 = P[id1, 1]  # 1/m^2
            if abs(k1) < kmin:
                k1 = 0
            EeV = 1e9 * energy  # eV
            brho = np.sqrt(EeV ** 2 - Er ** 2) / clight  # T-m
            G = brho * k1  # T/m
            GL = G * leng  # T
            if leng == 0:
                sname = 'kG2T'
                sval = 1 / T2kG
            else:
                sname = 'kG2T_Gdl2G'
                sval = 1 / (leng * T2kG)
            polarity = -np.sign(k1 + np.finfo(float).eps)  # add eps so that sign=1 when k1=0
            coorc = np.copy(coor[id1, :])  # m,rad
            QUAD.append({
                'idf': idf[id1],
                'id': idd[id1],
                'seq': seq[ids[id1]]['name'],
                'area': area[ida[id1]]['name'],
                'parent': area[ida[id1]]['parent'],
                'sector': FDN[id1].strip(),
                'xkey': kxn,
                'prim': kon,
                'name': mname,
                'type': T[id1].strip(),
                'dist': dist,
                'energy': energy,
                'leng': leng,
                'bore': bore,
                'tilt': np.rad2deg(tilt),  # deg
                'k1': k1,
                'GL': T2kG * GL,  # kG
                'G': charge * G,
                'sname': sname,
                'sval': sval,
                'polarity': polarity,
                'sdsp': sdsp,
                'suml': suml,
                **{f'c{k+1}': coorc[k] for k in range(6)},
                **{f'm{k+1}': [] for k in range(6)}
            })
    elif kwn == 'SEXT':
        id_ = id_[::2]
        name = [N[i] for i in id_]
        SEXT = []
        for mname in name:
            id_ = strmatch(mname,N,True)
            id1 = id_[0]  # first half (beam center)
            sdsp = Sd[id1]  # m
            suml = S[id1]  # m
            dist = suml - seq[ids[id1]]['suml']  # m
            energy = E[id1]  # GeV
            leng = np.sum(L[id_])  # m
            bore = 2 * A[id1]  # m
            tilt = P[id1, 3]  # rad
            k2 = P[id1, 2]  # 1/m^3
            if abs(k2) < kmin:
                k2 = 0
            EeV = 1e9 * energy  # eV
            brho = np.sqrt(EeV ** 2 - Er ** 2) / clight  # T-m
            Gp = brho * k2  # T/m
            GpL = Gp * leng  # T
            if leng == 0:
                sname = 'kG2T'
                sval = 1 / T2kG
            else:
                sname = 'kG2T_Gpdl2Gp'
                sval = 1 / (leng * T2kG)
            polarity = -np.sign(k2 + np.finfo(float).eps)  # add eps so that sign=1 when k2=0
            coorc = np.copy(coor[id1, :])  # m,rad
            SEXT.append({
                'idf': idf[id1],
                'id': idd[id1],
                'seq': seq[ids[id1]]['name'],
                'area': area[ida[id1]]['name'],
                'parent': area[ida[id1]]['parent'],
                'sector': FDN[id1].strip(),
                'xkey': kxn,
                'prim': kon,
                'name': mname,
                'type': T[id1].strip(),
                'dist': dist,
                'energy': energy,
                'leng': leng,
                'bore': bore,
                'tilt': np.rad2deg(tilt),  # deg
                'k2': k2,
                'GpL': T2kG * GpL,  # kG/m
                'Gp': charge * Gp,
                'sname': sname,
                'sval': sval,
                'polarity': polarity,
                'sdsp': sdsp,
                'suml': suml,
                **{f'c{k+1}': coorc[k] for k in range(6)}
            })
    elif kwn == 'SOLE':
        name = list(dict.fromkeys([N[i] for i in id_]))
        SOLE = []
        for mname in name:
            id_ = strmatch(mname,N,True)
            id1 = id_[0]
            ide = [id1 - 1, id_[-1]]  # [entrance, exit]
            sdsp = np.mean(Sd[ide])  # m
            suml = np.mean(S[ide])  # m
            dist = suml - seq[ids[id1]]['suml']  # m
            energy = np.mean(E[ide])  # GeV
            leng = np.sum(L[id_])  # m
            bore = 2 * A[id1]  # m
            ks = P[id1, 4]  # 1/m
            if abs(ks) < kmin:
                ks = 0
            EeV = 1e9 * energy  # eV
            brho = np.sqrt(EeV**2 - Er**2) / clight  # T-m
            B = brho * ks  # T
            BL = B * leng  # T-m
            if leng == 0:
                sname = 'kG2T'
                sval = 1 / T2kG
            else:
                sname = 'kG2T_Bdl2B'
                sval = 1 / (leng * T2kG)
            polarity = -np.sign(ks + np.finfo(float).eps)  # add eps so that sign=1 when ks=0
            coorc = np.mean(coor[ide], axis=0)  # m, rad
            SOLE.append({
                'idf': idf[id1],
                'id': idd[id1],
                'seq': seq[ids[id1]]['name'],
                'area': area[ida[id1]]['name'],
                'parent': area[ida[id1]]['parent'],
                'sector': FDN[id1].strip(),
                'xkey': kxn,
                'prim': kon,
                'name': mname,
                'type': T[id1].strip(),
                'dist': dist,
                'energy': energy,
                'leng': leng,
                'bore': bore,
                'ks': ks,
                'BL': T2kG * BL,  # kG-m
                'B': charge * B,
                'sname': sname,
                'sval': sval,
                'polarity': polarity,
                'sdsp': sdsp,
                'suml': suml,
                **{f'c{k+1}': coorc[k] for k in range(6)}
            })
    elif kwn == 'MATR':
        id_ = id_[::2]
        name = [N[i] for i in id_]
        MATR = []
        for mname in name:
            id_ = strmatch(mname,N,True)
            id1 = id_[0]  # first half (beam center)
            sdsp = Sd[id1]  # m
            suml = S[id1]  # m
            dist = suml - seq[ids[id1]]['suml']  # m
            energy = E[id1]  # GeV
            leng = np.sum(L[id_])  # m
            undl = P2[id1, 0]  # m
            undk = P2[id1, 1]  # 1
            coorc = np.copy(coor[id1])  # m, rad
            MATR.append({
                'idf': idf[id1],
                'id': idd[id1],
                'seq': seq[ids[id1]]['name'],
                'area': area[ida[id1]]['name'],
                'parent': area[ida[id1]]['parent'],
                'sector': FDN[id1].strip(),
                'xkey': kxn,
                'prim': kon,
                'name': mname,
                'type': T[id1].strip(),
                'dist': dist,
                'energy': energy,
                'leng': leng,
                'lambda': undl,
                'k': undk,
                'sdsp': sdsp,
                'suml': suml,
                **{f'c{k+1}': coorc[k] for k in range(6)}
            })
    elif kwn == 'RCOL':
        name = [N[i] for i in id_] # RCOLs are not split
        RCOL = []
        for mname in name:
            id_ = strmatch(mname,N,True)[0]
            ide = [id_ - 1, id_]  # [entrance, exit]
            sdsp = np.mean(Sd[ide])  # m (beam center)
            suml = np.mean(S[ide])  # m (beam center)
            dist = suml - seq[ids[id_]]['suml']  # m (sequence start to beam center)
            energy = E[id_]  # GeV
            leng = L[id_]  # m
            xgap = 2 * P[id_, 3]  # m
            ygap = 2 * P[id_, 4]  # m
            coorc = np.mean(coor[ide], axis=0)  # m, rad (beam center)
            RCOL.append({
                'idf': idf[id_],
                'id': idd[id_],
                'seq': seq[ids[id_]]['name'],
                'area': area[ida[id_]]['name'],
                'parent': area[ida[id_]]['parent'],
                'sector': FDN[id_].strip(),
                'xkey': kxn,
                'prim': kon,
                'name': mname,
                'type': T[id_].strip(),
                'dist': dist,
                'energy': energy,
                'leng': leng,
                'xgap': xgap,
                'ygap': ygap,
                'sdsp': sdsp,
                'suml': suml,
                **{f'c{k+1}': coorc[k] for k in range(6)}
            })
    elif kwn == 'ECOL':
        name = [N[i] for i in id_]  # ECOLs are not split
        ECOL = []
        for mname in name:
            id_ = strmatch(mname,N,True)[0]
            ide = [id_ - 1, id_]  # [entrance, exit]
            sdsp = np.mean(Sd[ide])  # m (beam center)
            suml = np.mean(S[ide])  # m (beam center)
            dist = suml - seq[ids[id_]]['suml']  # m (sequence start to beam center)
            energy = E[id_]  # GeV
            leng = L[id_]  # m
            xbore = 2 * P[id_, 3]  # m
            ybore = 2 * P[id_, 4]  # m
            coorc = np.mean(coor[ide], axis=0)  # m, rad (beam center)
            ECOL.append({
                'idf': idf[id_],
                'id': idd[id_],
                'seq': seq[ids[id_]]['name'],
                'area': area[ida[id_]]['name'],
                'parent': area[ida[id_]]['parent'],
                'sector': FDN[id_].strip(),
                'xkey': kxn,
                'prim': kon,
                'name': mname,
                'type': T[id_].strip(),
                'dist': dist,
                'energy': energy,
                'leng': leng,
                'xbore': xbore,
                'ybore': ybore,
                'sdsp': sdsp,
                'suml': suml,
                **{f'c{k+1}': coorc[k] for k in range(6)}
            })
    elif kwn == 'SROT':
        name = [N[i] for i in id_]  # SROTs are not split
        SROT = []
        for mname in name:
            id_ = strmatch(mname,N,True)[0]
            ide = [id_ - 1, id_]  # [entrance, exit]
            sdsp = np.mean(Sd[ide])  # m (beam center)
            suml = np.mean(S[ide])  # m (beam center)
            dist = suml - seq[ids[id_]]['suml']  # m (sequence start to beam center)
            energy = E[id_]  # GeV
            leng = L[id_]  # m
            ang = np.rad2deg(P[id_, 4])  # deg
            if abs(ang) < amin:
                ang = 0
            coorc = np.mean(coor[ide], axis=0)  # m, rad (beam center)
            SROT.append({
                'idf': idf[id_],
                'id': idd[id_],
                'seq': seq[ids[id_]]['name'],
                'area': area[ida[id_]]['name'],
                'parent': area[ida[id_]]['parent'],
                'sector': FDN[id_].strip(),
                'xkey': kxn,
                'prim': kon,
                'name': mname,
                'type': T[id_].strip(),
                'dist': dist,
                'energy': energy,
                'leng': leng,
                'ang': ang,
                'sdsp': sdsp,
                'suml': suml,
                **{f'c{k+1}': coorc[k] for k in range(6)}
            })
    elif kwn in [keyw[i] for i in idmisc]: #MISC
        name = [N[i] for i in id_]
        for mname in name:
            id_ = strmatch(mname,N,True)
            if id_[0] == 1:
                idi = 1
            else:
                idi = id_[0] - 1  # beam in
            ide = [idi, id_[-1]]  # [entrance, exit]
            sdsp = np.mean(Sd[ide])  # m (beam center)
            suml = np.mean(S[ide])  # m (beam center)
            dist = suml - seq[ids[id_[0]]]['suml']  # m (sequence start to beam center)
            energy = E[id_[0]]  # GeV
            leng = np.sum(L[id_])  # m
            coorc = np.mean(coor[ide], axis=0)  # m, rad (beam center)
            MISC = {
                'idf': idf[id_[0]],
                'id': idd[id_[0]],
                'seq': seq[ids[id_[0]]]['name'],
                'area': area[ida[id_[0]]]['name'],
                'parent': area[ida[id_[0]]]['parent'],
                'sector': FDN[id_[0]].strip(),
                'xkey': kxn,
                'prim': kon,
                'name': mname,
                'type': T[id_[0]].strip(),
                'dist': dist,
                'energy': energy,
                'leng': leng,
                'sdsp': sdsp,
                'suml': suml,
                **{f'c{k+1}': coorc[k] for k in range(6)}
            }
            globals()[f'n{kwn}'] += 1
            if f'{kwn}' in globals():
                globals()[f'{kwn}'].append(MISC)
            else:
                globals()[f'{kwn}'] = [MISC]
    elif kwn == 'MARK':
        name = [N[i] for i in id_]
        MARK = []
        for mname in name:
            id_ = strmatch(mname,N,True)[0]
            sdsp = Sd[id_]  # m
            suml = S[id_]  # m
            dist = suml - seq[ids[id_]]['suml']  # m (sequence start to beam center)
            energy = E[id_]  # GeV
            coorc = np.copy(coor[id_])
            MARK.append({
                'idf': idf[id_],
                'id': idd[id_],
                'seq': seq[ids[id_]]['name'],
                'area': area[ida[id_]]['name'],
                'parent': area[ida[id_]]['parent'],
                'sector': FDN[id_].strip(),
                'xkey': kxn,
                'prim': kon,
                'name': mname,
                'type': T[id_].strip(),
                'dist': dist,
                'energy': energy,
                'sdsp': sdsp,
                'suml': suml,
                **{f'c{k+1}': coorc[k] for k in range(6)}
            })

import scipy.io as sio

# deferred devices

DEPR = []
nDEPR = 0
aname = [x['name'] for x in area]
for k_str in keyw:
    k = globals()[k_str]
    for m in range(len(k)):
        t = k[m]['type']
        if t and t[0] == '@':
            deplev = int(t[1])
            if len(t) > 3:
                t = t[3:]  # skip ","
            else:
                t = ''
            nDEPR += 1
            id_ = strmatch(k[m]['area'],aname)
            if k == 'SBEN':
                z_use = k[m]['m1']
            else:
                z_use = k[m]['c1']
            DEPR.append({
                'id':k[m]['id'],
                'area':k[m]['area'],
                'ida': id_[0]+1,
                'prim': k[m]['prim'],
                'name': k[m]['name'],
                'z': z_use,
                'type': t,
                'level': deplev,
                })

            k[m]['area'] = '*' + k[m]['area']
            k[m]['type'] = t

# ------------------------------------------------------------------------------
# Fix magnet coordinates ...
# ------------------------------------------------------------------------------

def fix_magnet_coords(SBEN, QUAD, INST, K, N, L, P, coor):
  # Set special magnet coordinates
  bname = [x['name'] for x in SBEN]
 
  # center BC14 bends 1 and 4 on linac axis
  names=['BCX14720','BCX14883']
  for name in names:
    id_=strmatch(name,bname,True);
    SBEN[id_].m2=0
  
  # center first and last BC20 bends on linac axis
  names=['B1LE','B1RE']
  for name in names:
    id_=strmatch(name,bname,'exact')
    SBEN[id_].m2=0

  return SBEN, QUAD

SBEN, QUAD = fix_magnet_coords(SBEN, QUAD, INST, K, N, L, P, coor)

# ------------------------------------------------------------------------------

# Worksheets
# ==========
# one sheet per keyword
# one sheet for segments
# one sheet for sequences
# one sheet for deferred devices

# common worksheet header

head1 = ['MAD #','Sequence','Area','Sector','XAL Keyword', 
         'DB Keyword','MAD Name','Engineering Type','SeqDist','Energy']
head2 = ['Display S','SumL (linac)','X Coor (linac)','Y Coor (linac)', 
         'Z Coor (linac)','X Angle (linac)','Y Angle (linac)','Z Angle (linac)']
foot1=head1;
foot2 = ['Display S','SumL (linac)','MAD Z (linac)','MAD X (linac)',
         'MAD Y (linac)','MAD Psi (linac)','MAD Phi (linac)','MAD Theta (linac)']
unit1 = ['','','','','','','','','m','GeV']
unit2 = ['m','m','m','m','m','rad','rad','rad']

# Precision for coordinate output
prec = 1e-6

wb = pyxl.Workbook()
for keywn in keyw:
    thead2, tfoot2 = head2.copy(), foot2.copy()
    
    if keywn == 'LCAV':
      khead = ['Length', 'Frequency', 'Amplitude', 'Phase', 'Gradient', 'Power', 'Eloss']
      kunit = ['m', 'MHz', 'MeV', 'degree', 'MeV/m', 'fraction', 'V/C']
    elif keywn == 'SBEN':
      khead = ['Z Length', 'Effective Length', 'Gap', 'Field Integral', 'Tilt',
               'Angle', 'E1', 'E2', 'BL', 'B', 'K1', 'GL', 'G',
               'XAL Scale Name', 'XAL Scale Value', 'XAL Polarity']
      kunit = ['m', 'm', 'm', '', 'degree',
               'degree', 'degree', 'degree', 'kG-m', 'T', '1/m^2', 'kG', 'T/m',
               '', '', '']
      thead2.extend(['Magnet X Coor (linac)', 'Magnet Y Coor (linac)',
                     'Magnet Z Coor (linac)', 'Magnet X Angle (linac)',
                     'Magnet Y Angle (linac)', 'Magnet Z Angle (linac)'])
      tfoot2.extend(['Magnet MAD Z (linac)', 'Magnet MAD X (linac)',
                     'Magnet MAD Y (linac)', 'Magnet MAD Psi (linac)',
                     'Magnet MAD Phi (linac)', 'Magnet MAD Theta (linac)'])
      tunit2.extend(['m', 'm', 'm', 'rad', 'rad', 'rad'])
    elif keywn == 'QUAD':
      khead = ['Length','Bore','Tilt','K1','GL','G', 'XAL Scale Name',
               'XAL Scale Value','XAL Polarity']
      kunit = ['m','m','degree','1/m^2','kG','T/m', '','','']
      thead2.extend([ 'Magnet X Coor (linac)','Magnet Y Coor (linac)', 
                      'Magnet Z Coor (linac)','Magnet X Angle (linac)', 
                      'Magnet Y Angle (linac)','Magnet Z Angle (linac)'])
      tfoot2.extend(['Magnet MAD Z (linac)','Magnet MAD X (linac)', 
                     'Magnet MAD Y (linac)','Magnet MAD Psi (linac)', 
                     'Magnet MAD Phi (linac)','Magnet MAD Theta (linac)'])
      tunit2.extend(['m','m','m','rad','rad','rad'])
    elif keywn == 'SEXT':
      khead = ['Length','Bore','Tilt','K2','G''L','G''', 
               'XAL Scale Name','XAL Scale Value','XAL Polarity']
      kunit =['m','m','degree','1/m^3','kG/m','T/m^2','','','']
    elif keywn == 'SOLE':
      khead = ['Length','Bore','KS','BL','B',
               'XAL Scale Name','XAL Scale Value','XAL Polarity']
      kunit = ['m','m','1/m','kG-m','T','','','']
    elif keywn == 'MATR':
      khead = ['Length','Period','K']
      kunit = ['m','m','']
    elif keywn == 'RCOL':
      khead = ['Length','X Gap','Y Gap']
      kunit = ['m','m','m']
    elif keywn == 'ECOL':
      khead = ['Length','X Bore','Y Bore']
      kunit = ['m','m','m']
    elif keywn == 'SROT':
      khead = ['Length','Angle']
      kunit = ['m','deg']
    elif keywn in [MADK[x] for x in idmisc]:
      khead = ['Length']
      kunit = ['m']
    elif keywn == 'MARK':
      khead = ''
      kunit = ''
    
    kfoot = khead.copy()
    
    # Write to Excel file
    ws = wb.create_sheet(keywn)
    data = head1+khead+thead2
    for col, value in enumerate(data,start=1):
      ws.cell(row=1,column=col,value=value)
    data = unit1+kunit+tunit2
    for col, value in enumerate(data,start=1):
      ws.cell(row=2,column=col,value=value)
    
    # Process TEMP data
    M = []
    TEMP=globals()[keywn]
    for TEMPm in TEMP:
      for j in range(1, 7):
        TEMPm[f'c{j}'] = np.round(TEMPm[f'c{j}'], decimals=int(-np.log10(prec)))
        if TEMPm['prim'] == 'BEND' and TEMPm['type'] != '0.79K11.8':
          TEMPm[f'm{j}'] = np.round(TEMPm[f'm{j}'], decimals=int(-np.log10(prec)))
        elif TEMPm['prim'] == 'QUAD':
          TEMPm[f'm{j}'] = np.round(TEMPm[f'm{j}'], decimals=int(-np.log10(prec)))
        
        TEMPwr = {k: v for k, v in TEMPm.items() if k not in ['idf', 'area']}
        M.append(list(TEMPwr.values()))
    
    M.append(foot1 + kfoot + tfoot2)
    M.append(unit1 + kunit + tunit2)

    for i,datarow in enumerate(M,3):
      for j,data in enumerate(datarow,1):
        if j == 1 and isinstance(data,int):
          data = data + 1
        if isinstance(data,list):
          if data != []:
            ws.cell(row=i,column=j,value=data)
        elif isinstance(data,np.ndarray):
          if data.size > 0:
            ws.cell(row=i,column=j,value=data[0])
        else:
          ws.cell(row=i,column=j,value=data)

# ------------------------------------------------------------------------------
# Write Sequences Worksheet ...

# Sequences header

head = ['Sequence', 'Previous Sequence', 'Begin Element', 'End Element', 'Length']
unit = ['', '', '', '', 'm']

ws_seq = wb.create_sheet('Sequences')

for i,data in enumerate(head,1):
    ws_seq.cell(row=1,column=i,value=data)
for i,data in enumerate(unit,1):
    ws_seq.cell(row=2,column=i,value=data)

# Sequences Worksheet

M = []
for s in seq:
    temp = {
        'name': s['name'],
        'prev': '',
        'beg': s['beg'],
        'end': s['end'],
        'leng': s['length']
    }
    m = s['prev']
    if m > 0:
        temp['prev'] = seq[m-1]['name']
    M.append(list(temp.values()))

for i,row in enumerate(M,3):
    for j,data in enumerate(row,1):
        ws_seq.cell(row=i,column=j,value=data)

# ------------------------------------------------------------------------------
# Deferred devices header

head = ['MAD #', 'Area Name', 'Area Id', 'DB Keyword', 'MAD Name', 'Z',
        'Engineering Type', 'Level']

ws_def = wb.create_sheet('Deferred')
for i,data in enumerate(head,1):
    ws_def.cell(row=1,column=i,value=data)

# Deferred devices Worksheet

M = [list(DEPR[n].values()) for n in range(nDEPR)]
for i,row in enumerate(M,2):
    for j,data in enumerate(row,1):
        if j == 1 and isinstance(data,int):
            data = data + 1
        ws_def.cell(row=i,column=j,value=data)

Path(outdir).mkdir(parents=True,exist_ok=True)
wb.save(outdir+'/'+xfile)
# ------------------------------------------------------------------------------
# Write SYMBOLS txt-files ...

# set up pointers

seqname = [x['name'] for x in seq]
ip = []
for n in range(len(keyw)):
    TEMP = globals()[keyw[n]]
    for m in range(len(TEMP)):
        id_ = strmatch(TEMP[m]['seq'],seqname,True)[0]
        ip.append([seq[id_]['froot'], n, m, TEMP[m]['id']])
ip = sorted(ip, key=lambda x: (x[0], x[3]))


# SYMBOLS text-file headers and footers

head = ('Solid Edge,AREA,KeyW,ELEMENT,Eng_Name,L_EFF,'
        'APER,ANGLE,K1,K2,TILT,E1,E2,H1,H2,ENERGY,'
        'SUML,X Coor,Y Coor,Z Coor,X Angle,Y Angle,Z Angle,'
        'RF_Frequency,RF_Amplitude,RF_Phase,RF_Gradient,RF_Power_Fraction,'
        'Z_Length,Fringe_Field_Integral,Integrated_Field_BL,Field_B,'
        'Integrated_Field_Gradient_GL,Field_Gradient_G,'
        'XAL_Scale_Name,XAL_Scale_Value,XAL_Polarity,'
        'Magnet_X_Coor,Magnet_Y_Coor,Magnet_Z_Coor,'
        'Magnet_X_Angle,Magnet_Y_Angle,Magnet_Z_Angle,'
        'Solenoid_Strength_KS,Undulator_Period_Length,Undulator_Strength_K,'
        'X_Size,Y_Size,'
        'Section,Distance_From_Section_Start,XAL_Keyword,S_Display')

foot = ('MAD #,AREA,KeyW,ELEMENT,Eng_Name,L_EFF,'
        'APER,ANGLE,K1,K2,TILT,E1,E2,H1,H2,ENERGY,'
        'SUML,MAD Z,MAD X,MAD Y,MAD Psi,MAD Phi,MAD Theta,'
        'RF_Frequency,RF_Amplitude,RF_Phase,RF_Gradient,RF_Power_Fraction,'
        'Z_Length,Fringe_Field_Integral,Integrated_Field_BL,Field_B,'
        'Integrated_Field_Gradient_GL,Field_Gradient_G,'
        'XAL_Scale_Name,XAL_Scale_Value,XAL_Polarity,'
        'Magnet_MAD_Z,Magnet_MAD_X,Magnet_MAD_Y,'
        'Magnet_MAD_Psi,Magnet_MAD_Phi,Magnet_MAD_Theta,'
        'Solenoid_Strength_KS,Undulator_Period_Length,Undulator_Strength_K,'
        'X_Size,Y_Size,'
        'Section,Distance_From_Section_Start,XAL_Keyword,S_Display')

unit = (',,,,,m,m,'
        'm,deg,1/m^2,1/m^3,deg,deg,deg,1/m,1/m,GeV,'
        'm,m,m,m,rad,rad,rad,'
        'MHz,MeV,deg,MeV/m,1,'
        'm,1,kG-m,T,'
        'kG,T/m,'
        ',,,'
        'm,m,m,'
        'rad,rad,rad,'
        '1/m,m,1,'
        'm,m,'
        ',m,,m')

Ncol = head.count(',') + 1

# SYMBOLS text-file (linac coordinates)

def madval(rval):
    if rval == 0:
        return '0.0'
    else:
        aval = abs(rval)
        if 0.01 < aval < 100:
            iexp = 0.0
        else:
            iexp = int(math.log10(aval))
            rval = rval * 10**(-iexp)
        
        s = f'{rval:.12f}'
        s = s.rstrip('0')
        
        if s.endswith('.'):
            s = s + '0'
        
        if iexp != 0:
            s += f'E{iexp}'
        
        return s.strip()

def roundoff(val, prec=None):
    if isinstance(val,list):
        return None
    if prec is None:
        return val
    else:
        return prec * np.round(val / prec)

fname = f'FACET2-{optics}.txt'
with open(outdir+'/'+fname, 'wt') as fid:
    fid.write(f'{head}\n')
    fid.write(f'{unit}\n')

    for nf in range(1, len(froot) + 1):
        id_ = [i for i,x in enumerate(ip) if x[0]==nf]
        for n in id_:
            idk = ip[n][1]
            idn = ip[n][2]
            TEMP = globals()[keyw[idk]][idn]
            s = [None] * Ncol

            # common data

            s[0] = TEMP['id']
            s[1] = TEMP['area']
            s[2] = keyo[idk]
            s[3] = TEMP['name']
            s[4] = TEMP['type']
            s[15] = TEMP['energy']
            s[16] = TEMP['suml']
            s[17] = roundoff(TEMP['c1'], prec)
            s[18] = roundoff(TEMP['c2'], prec)
            s[19] = roundoff(TEMP['c3'], prec)
            s[20] = roundoff(TEMP['c4'], prec)
            s[21] = roundoff(TEMP['c5'], prec)
            s[22] = roundoff(TEMP['c6'], prec)
            s[48] = TEMP['seq']
            s[49] = TEMP['dist']
            s[50] = TEMP['xkey']
            s[51] = TEMP['sdsp']

            # keyword data
            if keyw[idk] == 'LCAV':
                s[5] = TEMP['leng']
                s[23] = TEMP['freq']
                s[24] = TEMP['ampl']
                s[25] = TEMP['phase']
                s[26] = TEMP['grad']
                s[27] = TEMP['power']
            elif keyw[idk] == 'SBEN':
                s[5] = TEMP['leng']
                s[6] = TEMP['gap']
                s[7] = TEMP['ang']
                s[8] = TEMP['k1']
                s[10] = TEMP['tilt']
                s[11] = TEMP['e1']
                s[12] = TEMP['e2']
                s[28] = TEMP['zleng']
                s[29] = TEMP['fint']
                s[30] = TEMP['BL']
                s[31] = TEMP['B']
                s[32] = TEMP['GL']
                s[33] = TEMP['G']
                s[34] = TEMP['sname']
                s[35] = TEMP['sval']
                s[36] = TEMP['polarity']
                s[37] = roundoff(TEMP['m1'], prec)
                s[38] = roundoff(TEMP['m2'], prec)
                s[39] = roundoff(TEMP['m3'], prec)
                s[40] = roundoff(TEMP['m4'], prec)
                s[41] = roundoff(TEMP['m5'], prec)
                s[42] = roundoff(TEMP['m6'], prec)
            elif keyw[idk] == 'QUAD':
                s[5] = TEMP['leng']
                s[6] = TEMP['bore']
                s[8] = TEMP['k1']
                s[10] = TEMP['tilt']
                s[32] = TEMP['GL']
                s[33] = TEMP['G']
                s[34] = TEMP['sname']
                s[35] = TEMP['sval']
                s[36] = TEMP['polarity']
                s[37] = roundoff(TEMP['m1'], prec)
                s[38] = roundoff(TEMP['m2'], prec)
                s[39] = roundoff(TEMP['m3'], prec)
                s[40] = roundoff(TEMP['m4'], prec)
                s[41] = roundoff(TEMP['m5'], prec)
                s[42] = roundoff(TEMP['m6'], prec)
            elif keyw[idk] == 'SEXT':
                s[5] = TEMP['leng']
                s[6] = TEMP['bore']
                s[9] = TEMP['k2']
                s[10] = TEMP['tilt']
                s[32] = TEMP['GpL']
                s[33] = TEMP['Gp']
                s[34] = TEMP['sname']
                s[35] = TEMP['sval']
                s[36] = TEMP['polarity']
            elif keyw[idk] == 'SOLE':
                s[5] = TEMP['leng']
                s[6] = TEMP['bore']
                s[30] = TEMP['BL']
                s[31] = TEMP['B']
                s[34] = TEMP['sname']
                s[35] = TEMP['sval']
                s[36] = TEMP['polarity']
                s[43] = TEMP['ks']
            elif keyw[idk] == 'MATR':
                s[5] = TEMP['leng']
                s[44] = TEMP['lambda']
                s[45] = TEMP['k']
            elif keyw[idk] == 'RCOL':
                s[5] = TEMP['leng']
                s[46] = TEMP['xgap']
                s[47] = TEMP['ygap']
            elif keyw[idk] == 'ECOL':
                s[5] = TEMP['leng']
                s[46] = TEMP['xbore']
                s[47] = TEMP['ybore']
            elif keyw[idk] == 'SROT':
                s[5] = TEMP['leng']
                s[7] = TEMP['ang']
            elif keyw[idk] in [keyw[i] for i in idmisc]:
                s[5] = TEMP['leng']

            fid.write(f"{s[0]+1},")
            for k in range(1, Ncol - 1):
                if s[k] is None:
                    fid.write(",")
                elif isinstance(s[k], np.ndarray):
                    fid.write(",")
                elif isinstance(s[k], str):
                    fid.write(f"{s[k]},")
                else:
                    fid.write(f"{madval(s[k])},")
            if isinstance(s[Ncol - 1], str):
                fid.write(f"{s[Ncol - 1]}\n")
            else:
                fid.write(f"{madval(s[Ncol - 1])}\n")
    
    fid.write(f'{foot}\n')
    fid.write(f'{unit}\n')

# ------------------------------------------------------------------------------
# Write extra SYMBOLS txt-file ...

# Element name, area name, undulator cell, sector

fname = f'FACET2-extra-{optics}.txt'
with open(outdir+'/'+fname, 'wt') as fid:
    fid.write('ELEMENT,Area2,,Sector\n')
    for nf in range(1,len(froot)+1):
        id_ = [i for i,x in enumerate(ip) if x[0]==nf]
        for n in id_:
            idk = ip[n][1]
            if keyw[idk] == 'MARK' or keyw[idk] == 'SROT':
                continue
            idn = ip[n][2]
            TEMP = globals()[keyw[idk]][idn]
            fid.write(f"{TEMP['name']},{TEMP['area']},,{TEMP['sector']}\n")
    fid.write('ELEMENT,Area2,,Sector\n')

# ------------------------------------------------------------------------------

# save RDBdata

